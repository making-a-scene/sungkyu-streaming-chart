import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DOWNLOAD_DIR = os.path.expanduser('~/Downloads')

import melon_all
import melon_30
import melon_100
import genie_all
import bugs
import flo
import vibe

CRAWLERS = [
    ('멜론 TOP 100', melon_all),
    ('멜론 HOT 100(30일)', melon_30),
    ('멜론 HOT 100(100일)', melon_100),
    ('지니 TOP 200', genie_all),
    ('벅스', bugs),
    ('플로', flo),
    ('바이브', vibe),
]

def crawl_all(artist='김성규'):
    results = []
    for chart_name, module in CRAWLERS:
        print(f"[크롤링 중] {chart_name}...")
        try:
            result = module.get_artist_rank(artist)
            results.append(result)

            if result['artist_ranks']:
                for item in result['artist_ranks']:
                    print(f"  -> {result['chart_name']} {item['rank']}위 - {item['title']}")
            else:
                print(f"  -> {result['chart_name']} -")
        except Exception as e:
            print(f"  -> {chart_name} 크롤링 실패: {e}")
            results.append({
                'chart_name': chart_name,
                'all_data': [],
                'artist_ranks': [],
            })
    return results


def export_to_excel(results, filename=None):
    if filename is None:
        today = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(DOWNLOAD_DIR, f'chart_{today}.xlsx')

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 1) 요약 시트
    ws_summary = wb.create_sheet('김성규 순위 요약')
    ws_summary.append(['차트', '순위', '곡명'])
    for col in range(1, 4):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 2
    for result in results:
        chart_name = result['chart_name']
        if result['artist_ranks']:
            for item in result['artist_ranks']:
                ws_summary.cell(row=row, column=1, value=chart_name).border = thin_border
                ws_summary.cell(row=row, column=2, value=item['rank']).border = thin_border
                ws_summary.cell(row=row, column=2).alignment = center_align
                ws_summary.cell(row=row, column=3, value=item['title']).border = thin_border
                row += 1
        else:
            ws_summary.cell(row=row, column=1, value=chart_name).border = thin_border
            ws_summary.cell(row=row, column=2, value='-').border = thin_border
            ws_summary.cell(row=row, column=2).alignment = center_align
            ws_summary.cell(row=row, column=3, value='-').border = thin_border
            row += 1

    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 40

    # 2) 각 차트별 전체 데이터 시트
    # for result in results:
    #     chart_name = result['chart_name']
    #     # 시트 이름은 31자 제한
    #     sheet_name = chart_name[:31]
    #     ws = wb.create_sheet(sheet_name)
    #     ws.append(['순위', '곡명', '아티스트'])
    #     for col in range(1, 4):
    #         cell = ws.cell(row=1, column=col)
    #         cell.font = header_font_white
    #         cell.fill = header_fill
    #         cell.alignment = center_align
    #         cell.border = thin_border
    #
    #     for item in result['all_data']:
    #         r = ws.max_row + 1
    #         ws.cell(row=r, column=1, value=item['rank']).alignment = center_align
    #         ws.cell(row=r, column=1).border = thin_border
    #         ws.cell(row=r, column=2, value=item['title']).border = thin_border
    #         ws.cell(row=r, column=3, value=item['artist']).border = thin_border
    #
    #     ws.column_dimensions['A'].width = 8
    #     ws.column_dimensions['B'].width = 40
    #     ws.column_dimensions['C'].width = 25

    wb.save(filename)
    print(f"\n엑셀 파일 저장 완료: {filename}")
    return filename


if __name__ == "__main__":
    print("=" * 50)
    print(f" 스트리밍 차트 크롤링 ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print("=" * 50)

    results = crawl_all()

    print("\n" + "=" * 50)
    print(" 엑셀 파일로 저장 중...")
    print("=" * 50)

    export_to_excel(results)
